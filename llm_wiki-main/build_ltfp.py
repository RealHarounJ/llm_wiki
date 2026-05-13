"""
Interactive Long-Term Financial Planning builder.
Run this script to add one row at a time as the user and professor discuss each step.
Call add_row(step) to insert the next row.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PATH = r'c:\Users\Haroun_Jaafar\Desktop\llm_wiki\LongTerm_Financial_Planning.xlsx'

DARK_BLUE  = "1F3864"; MID_BLUE = "2E75B6"; LIGHT_BLUE = "D6E4F0"
YELLOW = "FFF2CC";  GREEN = "E2EFDA"; RED = "FCE4D6"; WHITE = "FFFFFF"; GREY = "F2F2F2"
thin = Side(style="thin", color="BFBFBF"); thick = Side(style="medium", color="1F3864")
def tb():  return Border(left=thin,  right=thin,  top=thin,  bottom=thin)

EUR = '#,##0'; PCT = '0.00%'
YEARS = ["Actual 2021","2022","2023","2024","2025","2026"]

def load():
    return openpyxl.load_workbook(PATH)

def save(wb): wb.save(PATH)

def s(ws, row, col, val="", bold=False, bg=WHITE, fg="000000", align="left", fmt=None, size=11):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, color=fg, size=size)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = tb()
    if fmt: c.number_format = fmt
    return c

def cf(ws, row, col, val, bg=WHITE, bold=False, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.number_format = EUR
    c.font = Font(bold=bold, color=color, size=11)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = tb()

def build_blank():
    """Creates the blank shell with inputs and structure only."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Long-Term Financial Plan"

    widths = {"A":40,"B":16,"C":16,"D":16,"E":16,"F":16,"G":16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.row_dimensions[1].height = 36
    c = ws.cell(row=1, column=1, value="LONG-TERM FINANCIAL PLANNING — Build Together")
    c.font = Font(bold=True, color="FFFFFF", size=13)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=Side(style="medium",color="1F3864"),right=Side(style="medium",color="1F3864"),
                      top=Side(style="medium",color="1F3864"),bottom=Side(style="medium",color="1F3864"))
    ws.merge_cells("A1:G1")

    # Inputs
    ws.row_dimensions[2].height = 6
    sec_hdr = ws.cell(row=3, column=1, value="INPUT ASSUMPTIONS (Yellow = editable)")
    sec_hdr.font = Font(bold=True, color="FFFFFF", size=11)
    sec_hdr.fill = PatternFill("solid", fgColor=MID_BLUE)
    sec_hdr.alignment = Alignment(horizontal="left", vertical="center")
    sec_hdr.border = tb()
    ws.merge_cells("A3:G3")

    inputs = [
        (4,  "Revenue Growth Rate (g)",              0.20,   PCT),
        (5,  "Total Costs as % of Revenues",         0.92,   PCT),
        (6,  "Depreciation as % of Net Fixed Assets",0.09,   PCT),
        (7,  "Interest Rate on Long-term Debt",      0.085,  PCT),
        (8,  "Tax Rate",                             0.50,   PCT),
        (9,  "Dividend Payout Ratio (d)",            0.30,   PCT),
        (10, "Actual 2021 Revenues",                 2200,   EUR),
        (11, "Actual 2021 Long-term Debt",           757,    EUR),
        (12, "Actual 2021 Net Fixed Assets (NFA)",   268.5,  EUR),
        (13, "Actual 2021 Net Working Capital (NWC)",192,    EUR),
        (14, "Actual 2021 Shareholders Equity",      490,    EUR),
    ]
    for row, lbl, val, fmt in inputs:
        ws.row_dimensions[row].height = 22
        s(ws, row, 1, lbl, bg=LIGHT_BLUE)
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = fmt; c.font = Font(bold=True, size=11)
        c.fill = PatternFill("solid", fgColor=YELLOW)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = tb()

    # Year headers
    ws.row_dimensions[15].height = 6
    ws.row_dimensions[16].height = 26
    s(ws, 16, 1, "ITEM", bold=True, bg=DARK_BLUE, fg="FFFFFF", align="center")
    for i, y in enumerate(YEARS):
        c = ws.cell(row=16, column=2+i, value=y)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = tb()

    # Section headers (blank rows)
    ws.row_dimensions[17].height = 6
    ws.row_dimensions[18].height = 22
    sec = ws.cell(row=18, column=1, value="STEP 1 — Operating Cash Flow")
    sec.font = Font(bold=True, color="FFFFFF", size=11)
    sec.fill = PatternFill("solid", fgColor=MID_BLUE)
    sec.alignment = Alignment(horizontal="left", vertical="center"); sec.border = tb()
    ws.merge_cells("A18:G18")

    # Blank formula rows (labels only, formulas TBD)
    step1_labels = [
        (19, "Revenues"),
        (20, "Total Costs (92% of Revenues)"),
        (21, "EBIT (Revenues - Costs)"),
        (22, "Interest on Long-term Debt"),
        (23, "Income Before Tax"),
        (24, "Taxes"),
        (25, "Net Income"),
        (26, "+ Depreciation (add back)"),
        (27, "OPERATING CASH FLOW"),
    ]
    for row, lbl in step1_labels:
        ws.row_dimensions[row].height = 22
        is_total = lbl in ("EBIT (Revenues - Costs)","OPERATING CASH FLOW","Net Income","Income Before Tax")
        bg = GREEN if "OPERATING" in lbl else (LIGHT_BLUE if is_total else WHITE)
        s(ws, row, 1, lbl, bold=is_total or "OPERATING" in lbl, bg=bg)
        for j in range(6):
            c = ws.cell(row=row, column=2+j)
            c.fill = PatternFill("solid", fgColor=bg); c.border = tb()
            c.number_format = EUR

    ws.row_dimensions[28].height = 6
    sec2 = ws.cell(row=29, column=1, value="STEP 2 — Uses of Capital & EFN")
    sec2.font = Font(bold=True, color="FFFFFF", size=11)
    sec2.fill = PatternFill("solid", fgColor=MID_BLUE)
    sec2.alignment = Alignment(horizontal="left", vertical="center"); sec2.border = tb()
    ws.merge_cells("A29:G29")

    step2_labels = [
        (30, "Investment in NWC"),
        (31, "Gross Investment in Fixed Assets"),
        (32, "Dividends"),
        (33, "TOTAL USES OF CAPITAL"),
        (34, "EXTERNAL FINANCING NEEDED (EFN)"),
        (35, "Reinvested Earnings"),
    ]
    for row, lbl in step2_labels:
        ws.row_dimensions[row].height = 22
        bg = RED if "EFN" in lbl else (LIGHT_BLUE if "TOTAL" in lbl else WHITE)
        s(ws, row, 1, lbl, bold="TOTAL" in lbl or "EFN" in lbl, bg=bg)
        for j in range(6):
            c = ws.cell(row=row, column=2+j)
            c.fill = PatternFill("solid", fgColor=bg); c.border = tb()
            c.number_format = EUR

    ws.row_dimensions[36].height = 6
    sec3 = ws.cell(row=37, column=1, value="STEP 3 — Pro Forma Balance Sheet")
    sec3.font = Font(bold=True, color="FFFFFF", size=11)
    sec3.fill = PatternFill("solid", fgColor=MID_BLUE)
    sec3.alignment = Alignment(horizontal="left", vertical="center"); sec3.border = tb()
    ws.merge_cells("A37:G37")

    step3_labels = [
        (38, "ASSETS",                        DARK_BLUE, "FFFFFF"),
        (39, "Net Working Capital (NWC)",      WHITE,     "000000"),
        (40, "Net Fixed Assets (NFA)",         WHITE,     "000000"),
        (41, "TOTAL NET ASSETS",               GREEN,     DARK_BLUE),
        (42, "LIABILITIES & EQUITY",           DARK_BLUE, "FFFFFF"),
        (43, "Long-term Debt",                 WHITE,     "000000"),
        (44, "Shareholders Equity",            WHITE,     "000000"),
        (45, "TOTAL FINANCING",                GREEN,     DARK_BLUE),
        (46, "Balance Check (must = 0)",       GREY,      "C00000"),
    ]
    for row, lbl, bg, fg in step3_labels:
        ws.row_dimensions[row].height = 22
        bold = bg in (DARK_BLUE, GREEN, GREY)
        s(ws, row, 1, lbl, bold=bold, bg=bg, fg=fg)
        for j in range(6):
            c = ws.cell(row=row, column=2+j)
            c.fill = PatternFill("solid", fgColor=bg); c.border = tb()
            c.number_format = EUR

    ws.freeze_panes = "B17"
    ws.sheet_view.zoomScale = 90
    wb.save(PATH)
    print("Blank template saved!")

def add_revenues():
    wb = load(); ws = wb.active
    # B19 = Actual 2021
    cf(ws, 19, 2, "=B10")
    # C19 to G19 = prev * (1 + g)
    for j in range(1, 6):
        prev = get_column_letter(2+j-1)
        cf(ws, 19, 2+j, f"={prev}19*(1+$B$4)")
    save(wb)
    print("Row 19 (Revenues): DONE")
    print("  2022: =C19 -> 2,200 * 1.20 = 2,640")
    print("  2023: =D19 -> 2,640 * 1.20 = 3,168  ... and so on")

def add_costs():
    wb = load(); ws = wb.active
    for j in range(6):
        col = get_column_letter(2+j)
        cf(ws, 20, 2+j, f"=-{col}19*$B$5")
    save(wb)
    print("Row 20 (Total Costs): DONE")
    print("  2022: =-C19*0.92 -> -2,429")

def add_ebit():
    wb = load(); ws = wb.active
    for j in range(6):
        col = get_column_letter(2+j)
        c = ws.cell(row=21, column=2+j, value=f"={col}19+{col}20")
        c.number_format = EUR; c.font = Font(bold=True, color=DARK_BLUE)
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = tb()
    save(wb)
    print("Row 21 (EBIT): DONE")
    print("  2022: =C19+C20 -> 2,640 - 2,429 = 211")

def add_interest():
    wb = load(); ws = wb.active
    # B22 (Actual 2021): uses actual debt B11
    cf(ws, 22, 2, "=-$B$11*$B$7")
    # C22 onwards: uses prior year LT Debt from BS (row 43)
    for j in range(1, 6):
        prev = get_column_letter(2+j-1)
        cf(ws, 22, 2+j, f"=-{prev}43*$B$7")
    save(wb)
    print("Row 22 (Interest): DONE")
    print("  Actual 2021: =-757 * 0.085 = -64.3")
    print("  2022+: uses Long-term Debt from Balance Sheet (row 43)")

def add_ebt():
    wb = load(); ws = wb.active
    for j in range(6):
        col = get_column_letter(2+j)
        c = ws.cell(row=23, column=2+j, value=f"={col}21+{col}22")
        c.number_format = EUR; c.font = Font(bold=True, color=DARK_BLUE)
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = tb()
    save(wb)
    print("Row 23 (Income Before Tax): DONE")

def add_taxes():
    wb = load(); ws = wb.active
    for j in range(6):
        col = get_column_letter(2+j)
        cf(ws, 24, 2+j, f"=-{col}23*$B$8")
    save(wb)
    print("Row 24 (Taxes): DONE — NEGATIVE sign! Taxes are a cash outflow.")

def add_net_income():
    wb = load(); ws = wb.active
    for j in range(6):
        col = get_column_letter(2+j)
        c = ws.cell(row=25, column=2+j, value=f"={col}23+{col}24")
        c.number_format = EUR; c.font = Font(bold=True, color=DARK_BLUE)
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = tb()
    save(wb)
    print("Row 25 (Net Income): DONE")

def add_depreciation():
    wb = load(); ws = wb.active
    # B26 = actual NFA * depr rate
    cf(ws, 26, 2, "=$B$12*$B$6")
    # C26+ = prior year NFA (from BS row 40) * depr rate
    for j in range(1, 6):
        prev = get_column_letter(2+j-1)
        cf(ws, 26, 2+j, f"={prev}40*$B$6")
    save(wb)
    print("Row 26 (Depreciation): DONE")
    print("  Actual 2021: 268.5 * 9% = 24.2")
    print("  2022+: prior year NFA * 9%")

def add_ocf():
    wb = load(); ws = wb.active
    for j in range(6):
        col = get_column_letter(2+j)
        c = ws.cell(row=27, column=2+j, value=f"={col}25+{col}26")
        c.number_format = EUR; c.font = Font(bold=True, color=DARK_BLUE, size=12)
        c.fill = PatternFill("solid", fgColor=GREEN)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = tb()
    save(wb)
    print("Row 27 (OPERATING CASH FLOW): DONE")

def add_step2():
    wb = load(); ws = wb.active
    # Dividends (row 32)
    for j in range(6):
        col = get_column_letter(2+j)
        cf(ws, 32, 2+j, f"=-{col}25*$B$9")
    # Reinvested Earnings (row 35)
    for j in range(6):
        col = get_column_letter(2+j)
        cf(ws, 35, 2+j, f"={col}25*(1-$B$9)")
    save(wb)
    print("Row 32 (Dividends) and Row 35 (Reinvested Earnings): DONE")

def add_balance_sheet():
    wb = load(); ws = wb.active
    # NWC (row 39) — Actual 2021
    cf(ws, 39, 2, "=$B$13")
    # NFA (row 40) — Actual 2021
    cf(ws, 40, 2, "=$B$12")
    # LT Debt (row 43) — Actual 2021
    cf(ws, 43, 2, "=$B$11")
    # Equity (row 44) — Actual 2021
    cf(ws, 44, 2, "=$B$14")

    for j in range(1, 6):
        col  = get_column_letter(2+j)
        prev = get_column_letter(2+j-1)
        # NWC grows with revenues (% of revenues simplification)
        cf(ws, 39, 2+j, f"={prev}39*(1+$B$4)")
        # NFA = prior + gross investment - depreciation
        # gross investment estimated to keep % of revenues constant
        cf(ws, 40, 2+j, f"={prev}40*(1+$B$4)")
        # LT Debt: prior + EFN if negative (company borrows)
        cf(ws, 43, 2+j, f"={prev}43+MAX(0,-{col}34)")
        # Equity: prior + reinvested earnings
        cf(ws, 44, 2+j, f"={prev}44+{col}35")

    # Total Assets
    for j in range(6):
        col = get_column_letter(2+j)
        c = ws.cell(row=41, column=2+j, value=f"={col}39+{col}40")
        c.number_format = EUR; c.font = Font(bold=True, color=DARK_BLUE)
        c.fill = PatternFill("solid", fgColor=GREEN)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = tb()

    # Total Financing
    for j in range(6):
        col = get_column_letter(2+j)
        c = ws.cell(row=45, column=2+j, value=f"={col}43+{col}44")
        c.number_format = EUR; c.font = Font(bold=True, color=DARK_BLUE)
        c.fill = PatternFill("solid", fgColor=GREEN)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = tb()

    # Balance Check
    for j in range(6):
        col = get_column_letter(2+j)
        c = ws.cell(row=46, column=2+j, value=f"={col}41-{col}45")
        c.number_format = EUR; c.font = Font(bold=True, color="C00000")
        c.fill = PatternFill("solid", fgColor=GREY)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = tb()

    # Add NWC and FA investments to step 2 (rows 30, 31)
    for j in range(1, 6):
        col  = get_column_letter(2+j)
        prev = get_column_letter(2+j-1)
        cf(ws, 30, 2+j, f"=-({col}39-{prev}39)")
        cf(ws, 31, 2+j, f"=-({col}40-{prev}40+{col}26)")
    cf(ws, 30, 2, 0)
    cf(ws, 31, 2, f"=-30")

    # Total Uses
    for j in range(6):
        col = get_column_letter(2+j)
        c = ws.cell(row=33, column=2+j, value=f"={col}30+{col}31+{col}32")
        c.number_format = EUR; c.font = Font(bold=True, color=DARK_BLUE)
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = tb()

    # EFN
    for j in range(6):
        col = get_column_letter(2+j)
        c = ws.cell(row=34, column=2+j, value=f"={col}27+{col}33")
        c.number_format = EUR; c.font = Font(bold=True, color="C00000", size=12)
        c.fill = PatternFill("solid", fgColor=RED)
        c.alignment = Alignment(horizontal="right", vertical="center"); c.border = tb()

    save(wb)
    print("STEP 2 & STEP 3 (Balance Sheet): ALL DONE")

if __name__ == "__main__":
    print("Building blank template...")
    build_blank()
    print("\nAdding rows one by one...\n")
    add_revenues()
    add_costs()
    add_ebit()
    add_interest()
    add_ebt()
    add_taxes()
    add_net_income()
    add_depreciation()
    add_ocf()
    add_step2()
    add_balance_sheet()
    print("\nAll steps complete! Open the Excel file to review.")
