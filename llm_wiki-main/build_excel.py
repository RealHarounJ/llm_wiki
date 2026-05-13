import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Capital Budgeting Model"

# Colors
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
YELLOW     = "FFF2CC"
GREEN_FILL = "E2EFDA"
RED_FILL   = "FCE4D6"
WHITE      = "FFFFFF"
GREY       = "F2F2F2"

def fill(h): return PatternFill("solid", fgColor=h)
thin  = Side(style="thin",   color="BFBFBF")
thick = Side(style="medium", color="1F3864")
def tb():  return Border(left=thin,  right=thin,  top=thin,  bottom=thin)
def thb(): return Border(left=thick, right=thick, top=thick, bottom=thick)

col_widths = {"A":38,"B":16,"C":16,"D":16,"E":16,"F":16,"G":16,"H":16}
for c, w in col_widths.items():
    ws.column_dimensions[c].width = w

EUR = '#,##0.00 "€"'
PCT = "0.00%"

# ── helpers ─────────────────────────────────────────────────────────────────
def s(row, col, val="", bold=False, bg=WHITE, fg="000000",
      align="left", fmt=None, size=11, italic=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=bold, color=fg, size=size, italic=italic)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border    = tb()
    if fmt: c.number_format = fmt
    return c

def hdr(row, col, val, span=1, bg=DARK_BLUE):
    # unmerge first to avoid conflicts, then merge
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=True, color="FFFFFF", size=12)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = thb()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col+span-1)

def sec(row, col, val, span=8, bg=MID_BLUE):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=True, color="FFFFFF", size=11)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = tb()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col+span-1)

def inp(row, col, val, fmt="#,##0.00"):
    c = ws.cell(row=row, column=col, value=val)
    c.font          = Font(bold=True, color="000000", size=11)
    c.fill          = PatternFill("solid", fgColor=YELLOW)
    c.alignment     = Alignment(horizontal="right", vertical="center")
    c.border        = tb()
    c.number_format = fmt
    return c

def auto(row, col, formula, fmt=EUR, color="1F3864"):
    c = ws.cell(row=row, column=col, value=formula)
    c.number_format = fmt
    c.font          = Font(bold=True, color=color, size=11)
    c.fill          = PatternFill("solid", fgColor=GREEN_FILL)
    c.alignment     = Alignment(horizontal="right", vertical="center")
    c.border        = tb()

def cf(row, col, val, bg=WHITE, bold=False, color="000000"):
    c = ws.cell(row=row, column=col, value=val)
    c.font          = Font(bold=bold, color=color, size=11)
    c.fill          = PatternFill("solid", fgColor=bg)
    c.alignment     = Alignment(horizontal="right", vertical="center")
    c.border        = tb()
    c.number_format = EUR
    return c

# ════════════════════════════════════════════════════════════════════════════
# ROW 1 — Title
# ════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[1].height = 36
hdr(1, 1, "CAPITAL BUDGETING MODEL — Prof. Domenichelli", span=8)

# ════════════════════════════════════════════════════════════════════════════
# ROWS 3-14 — Input Assumptions
# ════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[2].height = 6
sec(3, 1, "INPUT ASSUMPTIONS  (Yellow = editable | Green = auto-computed)", span=8)

left_inputs = [
    (4,  "Initial Investment (Year 0)",    -1_200_000, EUR),
    (5,  "Useful Life (years)",                     6, "0"),
    (6,  "Annual Revenues",                   800_000, EUR),
    (7,  "Annual Operating Costs",           -350_000, EUR),
    (8,  "Cannibalization (lost sales/yr)",   -50_000, EUR),
    (9,  "Salvage Value (Year 6)",             300_000, EUR),
    (10, "Initial NWC Investment",             -80_000, EUR),
    (11, "Corporate Tax Rate",                    0.21, PCT),
    (12, "WACC",                                0.0888, PCT),
]
for row, lbl, val, fmt in left_inputs:
    ws.row_dimensions[row].height = 22
    s(row, 1, lbl, bg=LIGHT_BLUE)
    inp(row, 2, val, fmt)

right_inputs = [
    (4, "Equity (E)",          600_000, EUR),
    (5, "Debt (D)",            400_000, EUR),
    (6, "Cost of Equity rE",      0.12, PCT),
    (7, "Cost of Debt rD",        0.06, PCT),
    (8, "Tax Rate",               0.21, PCT),
]
for row, lbl, val, fmt in right_inputs:
    s(row, 5, lbl, bg=LIGHT_BLUE)
    inp(row, 6, val, fmt)

s(9,  5, "Depreciation / yr (auto)",   bg=LIGHT_BLUE); auto(9,  6, "=-B4/B5")
s(10, 5, "Book Value at Year 6",       bg=LIGHT_BLUE); auto(10, 6, "=B4+F9*B5")
s(11, 5, "Tax on Capital Gain (Yr 6)", bg=LIGHT_BLUE); auto(11, 6, "=(B9-F10)*B11")
s(12, 5, "Net Salvage CF (Yr 6)",      bg=LIGHT_BLUE); auto(12, 6, "=B9-F11")
s(13, 5, "WACC (computed)",            bg=LIGHT_BLUE); auto(13, 6, "=(F5/(F4+F5))*F7*(1-F8)+(F4/(F4+F5))*F6", fmt=PCT)

# ════════════════════════════════════════════════════════════════════════════
# ROW 16 — Year headers
# ════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[15].height = 6
ws.row_dimensions[16].height = 26
s(16, 1, "YEAR", bold=True, bg=DARK_BLUE, fg="FFFFFF", align="center")
for i, y in enumerate(["Year 0","Year 1","Year 2","Year 3","Year 4","Year 5","Year 6"]):
    c = ws.cell(row=16, column=2+i, value=y)
    c.font      = Font(bold=True, color="FFFFFF", size=11)
    c.fill      = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = tb()

# ════════════════════════════════════════════════════════════════════════════
# PANEL A — Capital Investment  (rows 17-20)
# ════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[17].height = 22
s(17, 1, "PANEL A — Cash Flow from Capital Investment",
  bold=True, bg=MID_BLUE, fg="FFFFFF")
for col in range(2, 9):
    c = ws.cell(row=17, column=col)
    c.fill = PatternFill("solid", fgColor=MID_BLUE); c.border = tb()

ws.row_dimensions[18].height = 22
s(18, 1, "Machine Purchase", bg=WHITE)
cf(18, 2, "=B4")
for j in range(1, 7): cf(18, 2+j, 0)

ws.row_dimensions[19].height = 22
s(19, 1, "Net Salvage Value (after-tax)", bg=WHITE)
for j in range(6): cf(19, 2+j, 0)
cf(19, 8, "=F12")

ws.row_dimensions[20].height = 22
s(20, 1, "TOTAL PANEL A", bold=True, bg=GREEN_FILL)
for j in range(7):
    col = get_column_letter(2+j)
    cf(20, 2+j, f"=SUM({col}18:{col}19)", bg=GREEN_FILL, bold=True, color="1F3864")

# ════════════════════════════════════════════════════════════════════════════
# PANEL B — Operating Cash Flow  (rows 22-34)
# ════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[21].height = 6
ws.row_dimensions[22].height = 22
s(22, 1, "PANEL B — Operating Cash Flow",
  bold=True, bg=MID_BLUE, fg="FFFFFF")
for col in range(2, 9):
    c = ws.cell(row=22, column=col)
    c.fill = PatternFill("solid", fgColor=MID_BLUE); c.border = tb()

b_rows = [
    (23, "Revenues",                WHITE),
    (24, "Cannibalization Effect",  WHITE),
    (25, "Operating Costs",         WHITE),
    (26, "EBITDA",                  LIGHT_BLUE),
    (27, "Depreciation",            WHITE),
    (28, "EBIT (Pre-tax Profit)",   LIGHT_BLUE),
    (29, "Taxes (21%)",             WHITE),
    (30, "Net Income",              LIGHT_BLUE),
    (31, "+ Add back Depreciation", WHITE),
    (32, "Operating Cash Flow",     GREEN_FILL),
]
for row, lbl, bg in b_rows:
    ws.row_dimensions[row].height = 22
    bold = bg in (LIGHT_BLUE, GREEN_FILL)
    s(row, 1, lbl, bold=bold, bg=bg)
    cf(row, 2, 0, bg=bg)  # Year 0

for j in range(6):
    col = get_column_letter(3+j)
    cf(23, 3+j, "=B6",                         bg=WHITE)
    cf(24, 3+j, "=B8",                         bg=WHITE)
    cf(25, 3+j, "=B7",                         bg=WHITE)
    cf(26, 3+j, f"={col}23+{col}24+{col}25",   bg=LIGHT_BLUE, bold=True, color="1F3864")
    cf(27, 3+j, "=F9",                         bg=WHITE)
    cf(28, 3+j, f"={col}26+{col}27",           bg=LIGHT_BLUE, bold=True, color="1F3864")
    cf(29, 3+j, f"=-{col}28*B11",              bg=WHITE)
    cf(30, 3+j, f"={col}28+{col}29",           bg=LIGHT_BLUE, bold=True, color="1F3864")
    cf(31, 3+j, f"=-{col}27",                  bg=WHITE)
    cf(32, 3+j, f"={col}30+{col}31",           bg=GREEN_FILL, bold=True, color="1F3864")

# ════════════════════════════════════════════════════════════════════════════
# PANEL C — Net Working Capital  (rows 34-36)
# ════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[33].height = 6
ws.row_dimensions[34].height = 22
s(34, 1, "PANEL C — Cash Flow from Net Working Capital",
  bold=True, bg=MID_BLUE, fg="FFFFFF")
for col in range(2, 9):
    c = ws.cell(row=34, column=col)
    c.fill = PatternFill("solid", fgColor=MID_BLUE); c.border = tb()

ws.row_dimensions[35].height = 22
s(35, 1, "Change in NWC", bg=WHITE)
cf(35, 2, "=B10")
for j in range(1, 6): cf(35, 3+j, 0)
cf(35, 8, "=-B10")

# ════════════════════════════════════════════════════════════════════════════
# TOTAL FREE CASH FLOW  (row 37)
# ════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[36].height = 6
ws.row_dimensions[37].height = 28
s(37, 1, "TOTAL FREE CASH FLOW  (A + B + C)",
  bold=True, bg=DARK_BLUE, fg="FFFFFF", size=12)
for j in range(7):
    col = get_column_letter(2+j)
    c = ws.cell(row=37, column=2+j, value=f"={col}20+{col}32+{col}35")
    c.number_format = EUR
    c.font          = Font(bold=True, color="FFFFFF", size=12)
    c.fill          = PatternFill("solid", fgColor=DARK_BLUE)
    c.alignment     = Alignment(horizontal="right", vertical="center")
    c.border        = tb()

# ════════════════════════════════════════════════════════════════════════════
# PROJECT VALUATION  (rows 39-46)
# ════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[38].height = 6
ws.row_dimensions[39].height = 22
s(39, 1, "PROJECT VALUATION", bold=True, bg=MID_BLUE, fg="FFFFFF")
for col in range(2, 9):
    c = ws.cell(row=39, column=col)
    c.fill = PatternFill("solid", fgColor=MID_BLUE); c.border = tb()

ws.row_dimensions[40].height = 22
s(40, 1, "Discount Factor  1 / (1+WACC)^t", bg=LIGHT_BLUE)
for j in range(7):
    c = ws.cell(row=40, column=2+j, value=f"=1/(1+B12)^{j}")
    c.number_format = "0.0000"
    c.font          = Font(color="1F3864")
    c.fill          = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.alignment     = Alignment(horizontal="right"); c.border = tb()

ws.row_dimensions[41].height = 22
s(41, 1, "Discounted Cash Flow (PV each year)", bg=LIGHT_BLUE)
for j in range(7):
    col = get_column_letter(2+j)
    c = ws.cell(row=41, column=2+j, value=f"={col}37*{col}40")
    c.number_format = EUR
    c.font          = Font(color="1F3864")
    c.fill          = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.alignment     = Alignment(horizontal="right"); c.border = tb()

# NPV
ws.row_dimensions[43].height = 30
s(43, 1, "NET PRESENT VALUE (NPV)", bold=True, bg=GREEN_FILL, size=12)
c = ws.cell(row=43, column=2, value="=NPV(B12,C37:H37)+B37")
c.number_format = EUR
c.font          = Font(bold=True, size=13, color="1F3864")
c.fill          = PatternFill("solid", fgColor=GREEN_FILL)
c.alignment     = Alignment(horizontal="right", vertical="center")
c.border        = tb()

s(43, 3, '=IF(B43>0,"ACCEPT - NPV positive","REJECT - NPV negative")',
  bold=True, bg=GREEN_FILL, fg="1F3864", size=12, align="left")
ws.merge_cells("C43:F43")

# IRR
ws.row_dimensions[44].height = 30
s(44, 1, "INTERNAL RATE OF RETURN (IRR)", bold=True, bg=RED_FILL, size=12)
c = ws.cell(row=44, column=2, value="=IRR(B37:H37)")
c.number_format = PCT
c.font          = Font(bold=True, size=13, color="C00000")
c.fill          = PatternFill("solid", fgColor=RED_FILL)
c.alignment     = Alignment(horizontal="right", vertical="center")
c.border        = tb()

s(44, 3, '=IF(B44>B12,"IRR exceeds WACC - ACCEPT","IRR below WACC - REJECT")',
  bold=True, bg=RED_FILL, fg="C00000", size=12, align="left")
ws.merge_cells("C44:F44")

# WACC reminder
ws.row_dimensions[45].height = 22
s(45, 1, "WACC (hurdle rate)", bg=GREY)
c = ws.cell(row=45, column=2, value="=B12")
c.number_format = PCT
c.font          = Font(bold=True, color="1F3864")
c.fill          = PatternFill("solid", fgColor=GREY)
c.alignment     = Alignment(horizontal="right"); c.border = tb()

# ════════════════════════════════════════════════════════════════════════════
# Legend
# ════════════════════════════════════════════════════════════════════════════
ws.row_dimensions[47].height = 6
ws.row_dimensions[48].height = 20
s(48, 1, "LEGEND", bold=True, bg=MID_BLUE, fg="FFFFFF")
for col in range(2, 9):
    c = ws.cell(row=48, column=col)
    c.fill = PatternFill("solid", fgColor=MID_BLUE); c.border = tb()

ws.row_dimensions[49].height = 20
s(49, 1, "Yellow = INPUT cells (modify these)", bold=True, bg=YELLOW)
ws.merge_cells("A49:C49")
s(49, 4, "Green = AUTO-COMPUTED output",        bold=True, bg=GREEN_FILL)
ws.merge_cells("D49:F49")
s(49, 7, "All formulas update in real time",    bold=True, bg=LIGHT_BLUE)
ws.merge_cells("G49:H49")

# Freeze & zoom
ws.freeze_panes   = "B17"
ws.sheet_view.zoomScale = 90

path = r"c:\Users\Haroun_Jaafar\Desktop\llm_wiki\Capital_Budgeting_Model.xlsx"
wb.save(path)
print("Saved to", path)
