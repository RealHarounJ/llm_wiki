import openpyxl

wb = openpyxl.load_workbook(
    r'c:\Users\Haroun_Jaafar\Desktop\llm_wiki\Capital_Budgeting_Model.xlsx',
    data_only=False
)
ws = wb.active

print("=== CAPITAL BUDGETING MODEL — AUDIT REPORT ===\n")

def chk(label, cell_ref, expected=None):
    val = ws[cell_ref].value
    status = ""
    if expected is not None:
        status = " ✅" if str(val) == str(expected) else f" ⚠️  (expected: {expected})"
    print(f"  {label:35s} {str(val):50s}{status}")

print("--- INPUT ASSUMPTIONS ---")
chk("B4  Initial Investment",     "B4",  -1200000)
chk("B5  Useful Life",            "B5",  6)
chk("B6  Annual Revenues",        "B6",  800000)
chk("B7  Operating Costs",        "B7",  -350000)
chk("B8  Cannibalization",        "B8",  -50000)
chk("B9  Salvage Value",          "B9",  300000)
chk("B10 Initial NWC",            "B10", -80000)
chk("B11 Tax Rate",               "B11", 0.21)
chk("B12 WACC (manual)",          "B12", 0.0888)

print("\n--- AUTO-COMPUTED FORMULAS ---")
chk("F9  Depreciation/yr",        "F9",  "=-B4/B5")
chk("F10 Book Value at Year 6",   "F10", "=B4+F9*B5")
chk("F11 Tax on Capital Gain",    "F11", "=(B9-F10)*B11")
chk("F12 Net Salvage CF",         "F12", "=B9-F11")
chk("F13 WACC computed",          "F13", "=(F5/(F4+F5))*F7*(1-F8)+(F4/(F4+F5))*F6")

print("\n--- PANEL A: Capital Investment ---")
for row, lbl in [(18,"Machine Purchase"), (19,"Net Salvage"), (20,"TOTAL PANEL A")]:
    vals = [ws.cell(row=row, column=c).value for c in range(2, 9)]
    print(f"  Row {row} [{lbl}]")
    for i, v in enumerate(vals):
        print(f"    Year {i}: {v}")

print("\n--- PANEL B: Operating Cash Flow ---")
b_rows = {
    23:"Revenues", 24:"Cannibalization", 25:"Op Costs",
    26:"EBITDA", 27:"Depreciation", 28:"EBIT",
    29:"Taxes", 30:"Net Income", 31:"Add back Depr", 32:"Op CF"
}
for row, lbl in b_rows.items():
    vals = [ws.cell(row=row, column=c).value for c in range(2, 9)]
    print(f"  Row {row} [{lbl}]:")
    for i, v in enumerate(vals):
        print(f"    Year {i}: {v}")

print("\n--- PANEL C: NWC ---")
vals = [ws.cell(row=35, column=c).value for c in range(2, 9)]
for i, v in enumerate(vals):
    print(f"  Year {i}: {v}")

print("\n--- TOTAL FREE CASH FLOW ---")
vals = [ws.cell(row=37, column=c).value for c in range(2, 9)]
for i, v in enumerate(vals):
    print(f"  Year {i}: {v}")

print("\n--- VALUATION ---")
chk("NPV formula (B43)",  "B43", "=NPV(B12,C37:H37)+B37")
chk("IRR formula (B44)",  "B44", "=IRR(B37:H37)")
chk("WACC ref   (B45)",   "B45", "=B12")

print("\n=== AUDIT COMPLETE ===")
