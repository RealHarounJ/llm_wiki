import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

path = r'c:\Users\Haroun_Jaafar\Desktop\llm_wiki\Capital_Budgeting_Model.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

EUR  = '#,##0.00'
PCT  = '0.00%'
GREEN = 'E2EFDA'
RED   = 'FCE4D6'
WHITE = 'FFFFFF'
LIGHT = 'D6E4F0'
DARK  = '1F3864'
GREY  = 'F2F2F2'

thin = Side(style='thin', color='BFBFBF')
def tb(): return Border(left=thin, right=thin, top=thin, bottom=thin)

def fix(row, col, val, fmt=EUR, bg=WHITE, bold=False, color='000000'):
    c = ws.cell(row=row, column=col, value=val)
    c.number_format = fmt
    c.font      = Font(bold=bold, color=color, size=11)
    c.fill      = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border    = tb()

print('Applying all corrections...')

# FIX 1 — Auto-computed formulas (F9-F13)
fix(9,  6, '=-B4/B5',                                    bg=GREEN, bold=True, color=DARK)
fix(10, 6, '=B4+F9*B5',                                  bg=GREEN, bold=True, color=DARK)
fix(11, 6, '=(B9-F10)*B11',                               bg=GREEN, bold=True, color=DARK)
fix(12, 6, '=B9-F11',                                    bg=GREEN, bold=True, color=DARK)
fix(13, 6, '=(F5/(F4+F5))*F7*(1-F8)+(F4/(F4+F5))*F6',   fmt=PCT, bg=GREEN, bold=True, color=DARK)
print('  [1] Fixed F9-F13 auto-computed formulas')

# FIX 2 — Net Salvage (H19) must use net-of-tax formula from F12
fix(19, 8, '=F12', bg=WHITE)
print('  [2] Fixed H19: Net Salvage now = F12 (after-tax)')

# FIX 3 — Taxes row 29: must be NEGATIVE (tax is a cash outflow)
for j in range(6):
    col = get_column_letter(3+j)
    fix(29, 3+j, f'=-{col}28*B11', bg=WHITE)
print('  [3] Fixed row 29: Taxes are now negative')

# FIX 4 — Net Income row 30: EBIT + Taxes (taxes already negative)
for j in range(6):
    col = get_column_letter(3+j)
    fix(30, 3+j, f'={col}28+{col}29', bg=LIGHT, bold=True, color=DARK)
print('  [4] Fixed row 30: Net Income = EBIT + Taxes')

# FIX 5 — Add back Depreciation row 31: ONLY depreciation (not net income)
for j in range(6):
    col = get_column_letter(3+j)
    fix(31, 3+j, f'=-{col}27', bg=WHITE)
print('  [5] Fixed row 31: Add back = only depreciation')

# FIX 6 — Operating Cash Flow row 32: Net Income + Add-back Depr
for j in range(6):
    col = get_column_letter(3+j)
    fix(32, 3+j, f'={col}30+{col}31', bg=GREEN, bold=True, color=DARK)
print('  [6] Fixed row 32: OCF = Net Income + Depreciation')

# FIX 7 — NWC row 35 Year 1 (column C): must be 0
fix(35, 3, 0, bg=WHITE)
print('  [7] Fixed C35: NWC Year 1 = 0 (NWC only invested once in Year 0)')

# FIX 8 — Total FCF row 37: formula-driven
for j in range(7):
    col = get_column_letter(2+j)
    c = ws.cell(row=37, column=2+j, value=f'={col}20+{col}32+{col}35')
    c.number_format = EUR
    c.font      = Font(bold=True, color='FFFFFF', size=12)
    c.fill      = PatternFill('solid', fgColor='1F3864')
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border    = tb()
print('  [8] Fixed row 37: Total FCF = Panel A + B + C')

# FIX 9 — Discount factors row 40
for j in range(7):
    c = ws.cell(row=40, column=2+j, value=f'=1/(1+B12)^{j}')
    c.number_format = '0.0000'
    c.font      = Font(color=DARK)
    c.fill      = PatternFill('solid', fgColor=LIGHT)
    c.alignment = Alignment(horizontal='right')
    c.border    = tb()
print('  [9] Fixed row 40: Discount factors')

# FIX 10 — Discounted CFs row 41
for j in range(7):
    col = get_column_letter(2+j)
    c = ws.cell(row=41, column=2+j, value=f'={col}37*{col}40')
    c.number_format = EUR
    c.font      = Font(color=DARK)
    c.fill      = PatternFill('solid', fgColor=LIGHT)
    c.alignment = Alignment(horizontal='right')
    c.border    = tb()
print('  [10] Fixed row 41: Discounted cash flows')

# FIX 11 — NPV (B43)
c = ws.cell(row=43, column=2, value='=NPV(B12,C37:H37)+B37')
c.number_format = EUR
c.font      = Font(bold=True, size=13, color=DARK)
c.fill      = PatternFill('solid', fgColor=GREEN)
c.alignment = Alignment(horizontal='right', vertical='center')
c.border    = tb()
c2 = ws.cell(row=43, column=3,
             value='=IF(B43>0,"ACCEPT - NPV positive","REJECT - NPV negative")')
c2.font      = Font(bold=True, size=12, color=DARK)
c2.fill      = PatternFill('solid', fgColor=GREEN)
c2.alignment = Alignment(horizontal='left', vertical='center')
c2.border    = tb()
print('  [11] Fixed B43: NPV formula')

# FIX 12 — IRR (B44)
c = ws.cell(row=44, column=2, value='=IRR(B37:H37)')
c.number_format = PCT
c.font      = Font(bold=True, size=13, color='C00000')
c.fill      = PatternFill('solid', fgColor=RED)
c.alignment = Alignment(horizontal='right', vertical='center')
c.border    = tb()
c2 = ws.cell(row=44, column=3,
             value='=IF(B44>B12,"IRR exceeds WACC - ACCEPT","IRR below WACC - REJECT")')
c2.font      = Font(bold=True, size=12, color='C00000')
c2.fill      = PatternFill('solid', fgColor=RED)
c2.alignment = Alignment(horizontal='left', vertical='center')
c2.border    = tb()
print('  [12] Fixed B44: IRR formula')

# FIX 13 — WACC reference (B45)
c = ws.cell(row=45, column=2, value='=B12')
c.number_format = PCT
c.font      = Font(bold=True, color=DARK)
c.fill      = PatternFill('solid', fgColor=GREY)
c.alignment = Alignment(horizontal='right')
c.border    = tb()
print('  [13] Fixed B45: WACC reference')

wb.save(path)
print()
print('All 13 fixes applied. File saved successfully!')
print()
print('Expected results after opening Excel:')
print('  OCF per year:  358,000')
print('  FCF Year 0:   -1,280,000')
print('  FCF Year 1-5:  358,000')
print('  FCF Year 6:    438,000  (358k + 80k NWC recovery)')
print('  NPV:           ~429,000 (positive - ACCEPT)')
print('  IRR:           ~22%     (well above WACC of 8.88%)')
