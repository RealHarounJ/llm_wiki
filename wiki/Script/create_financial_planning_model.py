import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_planning_model():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Planning Model"
    
    # Attiva le linee della griglia (gridlines)
    ws.views.sheetView[0].showGridLines = True
    
    # Stili e Palette Colori (Sleek Professional Dark Navy)
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1B365D")
    section_font = Font(name=font_family, size=12, bold=True, color="1B365D")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=11)
    italic_font = Font(name=font_family, size=9, italic=True, color="555555")
    
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    summary_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    alert_fill = PatternFill(start_color="FDF2F2", end_color="FDF2F2", fill_type="solid")
    
    thin_side = Side(border_style="thin", color="D3D3D3")
    double_side = Side(border_style="double", color="1B365D")
    thick_bottom = Side(border_style="medium", color="1B365D")
    
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    bottom_double_border = Border(top=thin_side, bottom=double_side)
    header_border = Border(bottom=thick_bottom, left=thin_side, right=thin_side)
    
    # 1. Titolo del Foglio
    ws['A1'] = "DYNAMIC LONG-TERM FINANCIAL PLANNING MODEL"
    ws['A1'].font = title_font
    ws['A2'] = "Pro-Forma Three-Statement Integration & EFN Calculator (Excel Style with Real Formulas)"
    ws['A2'].font = italic_font
    
    # 2. Input del Modello (Sezione Parametri)
    ws['A4'] = "MODEL PARAMETERS & DRIVERS"
    ws['A4'].font = section_font
    
    ws['A5'] = "Sales Growth Rate (g)"
    ws['B5'] = 0.20
    ws['B5'].number_format = '0.0%'
    
    ws['A6'] = "Net Profit Margin (PM)"
    ws['B6'] = 0.10
    ws['B6'].number_format = '0.0%'
    
    ws['A7'] = "Dividend Payout Ratio (d)"
    ws['B7'] = 0.40
    ws['B7'].number_format = '0.0%'
    
    ws['A8'] = "Retention Ratio (b = 1 - d)"
    ws['B8'] = "=1-B7"
    ws['B8'].number_format = '0.0%'
    
    for row in range(5, 9):
        ws[f'A{row}'].font = regular_font
        ws[f'B{row}'].font = bold_font
        ws[f'A{row}'].border = thin_border
        ws[f'B{row}'].border = thin_border
        ws[f'B{row}'].alignment = Alignment(horizontal="right")
        
    # 3. Tabella Principale: Pro-Forma Income Statement & Balance Sheet
    ws['A10'] = "PRO-FORMA FINANCIAL STATEMENTS"
    ws['A10'].font = section_font
    
    headers = ["Financial Statement Item", "Current Year (t0)", "% of Sales (t0)", "Pro-Forma Year (t1)", "Excel Formula Used"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center", wrap_text=True)
        cell.border = header_border
    
    # Dati del Conto Economico (Income Statement)
    ws['A12'] = "INCOME STATEMENT"
    ws['A12'].font = bold_font
    
    ws['A13'] = "Sales (Vendite)"
    ws['B13'] = 1000.0
    ws['D13'] = "=B13*(1+$B$5)"
    
    ws['A14'] = "Net Income (Utile Netto)"
    ws['B14'] = "=B13*$B$6"
    ws['D14'] = "=D13*$B$6"
    
    ws['A15'] = "Dividends Paid (Dividendi)"
    ws['B15'] = "=B14*$B$7"
    ws['D15'] = "=D14*$B$7"
    
    ws['A16'] = "Retained Earnings of the Year"
    ws['B16'] = "=B14*$B$8"
    ws['D16'] = "=D14*$B$8"
    
    # Dati dello Stato Patrimoniale (Balance Sheet)
    ws['A18'] = "BALANCE SHEET"
    ws['A18'].font = bold_font
    
    # Asset (Attività)
    ws['A19'] = "ASSETS (Attività che crescono con le vendite)"
    ws['A19'].font = italic_font
    
    ws['A20'] = "Cash & Marketable Securities"
    ws['B20'] = 50.0
    ws['C20'] = "=B20/$B$13"
    ws['D20'] = "=D13*C20"
    
    ws['A21'] = "Accounts Receivable (Crediti)"
    ws['B21'] = 150.0
    ws['C21'] = "=B21/$B$13"
    ws['D21'] = "=D13*C21"
    
    ws['A22'] = "Inventory (Magazzino)"
    ws['B22'] = 200.0
    ws['C22'] = "=B22/$B$13"
    ws['D22'] = "=D13*C22"
    
    ws['A23'] = "Net Fixed Assets (Immobilizzazioni)"
    ws['B23'] = 600.0
    ws['C23'] = "=B23/$B$13"
    ws['D23'] = "=D13*C23"
    
    ws['A24'] = "TOTAL ASSETS (Attività Totali)"
    ws['B24'] = "=SUM(B20:B23)"
    ws['D24'] = "=SUM(D20:D23)"
    ws['B24'].font = bold_font
    ws['D24'].font = bold_font
    ws['B24'].fill = summary_fill
    ws['D24'].fill = summary_fill
    ws['B24'].border = bottom_double_border
    ws['D24'].border = bottom_double_border
    
    # Passività e Equity
    ws['A26'] = "LIABILITIES & EQUITY (Passività e Patrimonio)"
    ws['A26'].font = italic_font
    
    ws['A27'] = "Accounts Payable (Debiti Commerciali Spontanei)"
    ws['B27'] = 100.0
    ws['C27'] = "=B27/$B$13"
    ws['D27'] = "=D13*C27"
    
    ws['A28'] = "Long-Term Debt (Debito a L/T - Fisso)"
    ws['B28'] = 400.0
    ws['D28'] = "=B28"
    
    ws['A29'] = "Common Stock (Capitale Sociale - Fisso)"
    ws['B29'] = 300.0
    ws['D29'] = "=B29"
    
    ws['A30'] = "Accumulated Retained Earnings (Utili Trattenuti)"
    ws['B30'] = 200.0
    ws['D30'] = "=B30+D16"
    
    ws['A31'] = "TOTAL LIABILITIES & EQUITY (Prima dell'EFN)"
    ws['B31'] = "=SUM(B27:B30)"
    ws['D31'] = "=SUM(D27:D30)"
    ws['B31'].font = bold_font
    ws['D31'].font = bold_font
    ws['B31'].fill = summary_fill
    ws['D31'].fill = summary_fill
    ws['B31'].border = bottom_double_border
    ws['D31'].border = bottom_double_border
    
    # 4. Sezione EFN (Il Buco di Bilancio da quadrare)
    ws['A33'] = "EXTERNAL FUNDS NEEDED (EFN)"
    ws['A33'].font = section_font
    
    ws['A34'] = "EFN (Balance Sheet Discrepancy: Assets - Liabilities/Equity)"
    ws['B34'] = "=B24-B31"
    ws['D34'] = "=D24-D31"
    ws['B34'].font = bold_font
    ws['D34'].font = bold_font
    ws['B34'].fill = alert_fill
    ws['D34'].fill = alert_fill
    ws['B34'].border = thin_border
    ws['D34'].border = thin_border
    ws['B34'].alignment = Alignment(horizontal="right")
    ws['D34'].alignment = Alignment(horizontal="right")
    
    # 5. Tabella Fonti e Impieghi (Sources & Uses of Cash - Incremental Approach)
    ws['A36'] = "SOURCES & USES OF CASH (Pro-Forma Cash Flow Statement - Incremental Approach)"
    ws['A36'].font = section_font
    
    ws.merge_cells('A37:D37')
    ws['A37'] = "Questo prospetto ricava i flussi basandosi ESCLUSIVAMENTE sulle variazioni incrementali (t1 - t0) dei saldi dello Stato Patrimoniale."
    ws['A37'].font = italic_font
    
    headers_cf = ["Cash Flow Item (Fonti e Impieghi)", "Incremental Formula", "Cash Impact (€)", "Flow Type"]
    for col_idx, header in enumerate(headers_cf, 1):
        cell = ws.cell(row=38, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
        cell.border = header_border
        
    # Righe di calcolo flussi
    ws['A39'] = "Operating Cash Inflow: Net Income (Utile netto)"
    ws['B39'] = "Income Statement D14"
    ws['C39'] = "=D14"
    ws['D39'] = "Source"
    
    ws['A40'] = "Increase in Payables (Aumento debiti commerciali)"
    ws['B40'] = "D27 - B27"
    ws['C40'] = "=D27-B27"
    ws['D40'] = "Source (Liab. Increase)"
    
    ws['A41'] = "Increase in Cash Reserve (Aumento cassa operativa)"
    ws['B41'] = "-(D20 - B20)"
    ws['C41'] = "=-(D20-B20)"
    ws['D41'] = "Use (Asset Increase)"
    
    ws['A42'] = "Increase in Accounts Receivable (Aumento crediti)"
    ws['B42'] = "-(D21 - B21)"
    ws['C42'] = "=-(D21-B21)"
    ws['D42'] = "Use (Asset Increase)"
    
    ws['A43'] = "Increase in Inventory (Aumento magazzino)"
    ws['B43'] = "-(D22 - B22)"
    ws['C43'] = "=-(D22-B22)"
    ws['D43'] = "Use (Asset Increase)"
    
    ws['A44'] = "Increase in Net Fixed Assets (CapEx / Macchinari)"
    ws['B44'] = "-(D23 - B23)"
    ws['C44'] = "=-(D23-B23)"
    ws['D44'] = "Use (Asset Increase)"
    
    ws['A45'] = "Dividends Paid (Dividendi distribuiti)"
    ws['B45'] = "-D15"
    ws['C45'] = "=-D15"
    ws['D45'] = "Use (Outflow)"
    
    ws['A46'] = "NET CASH FLOW SURPLUS / (DEFICIT)"
    ws['B46'] = "SUM(C39:C45)"
    ws['C46'] = "=SUM(C39:C45)"
    ws['D46'] = "EFN Gap"
    ws['A46'].font = bold_font
    ws['C46'].font = bold_font
    ws['A46'].fill = alert_fill
    ws['C46'].fill = alert_fill
    ws['A46'].border = thin_border
    ws['C46'].border = thin_border
    ws['C46'].alignment = Alignment(horizontal="right")
    
    # 6. Copertura EFN (Plug Variable)
    ws['A48'] = "FINANCING OF THE EFN (Closing the balance sheet gap)"
    ws['A48'].font = bold_font
    
    ws['A49'] = "New External Debt Issued (Nuova emissione debito)"
    ws['C49'] = "=IF(C46<0, -C46*0.50, 0)"
    ws['D49'] = "Plug (50% of Deficit)"
    
    ws['A50'] = "New External Common Stock Issued (Nuove azioni)"
    ws['C50'] = "=IF(C46<0, -C46*0.50, 0)"
    ws['D50'] = "Plug (50% of Deficit)"
    
    ws['A51'] = "RECONCILED FINAL BALANCE (Net Cash Flow + Financing)"
    ws['C51'] = "=C46+C49+C50"
    ws['D51'] = "Reconciled to 0"
    ws['A51'].font = bold_font
    ws['C51'].font = bold_font
    ws['A51'].fill = summary_fill
    ws['C51'].fill = summary_fill
    ws['C51'].border = bottom_double_border
    ws['C51'].alignment = Alignment(horizontal="right")
    
    # Formattazione Numerica Generica di tutta la griglia
    for row in range(12, 52):
        for col in [2, 3, 4]:
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                # Se è percentuale
                if col == 3 and row >= 20 and row <= 27:
                    cell.number_format = '0.0%'
                elif isinstance(cell.value, float) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                    cell.number_format = '€#,##0.00'
                cell.alignment = Alignment(horizontal="right")
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            if row not in [24, 31, 34, 46, 51]: # non sovrascrivere bordi speciali
                cell.border = Border(bottom=thin_side, left=thin_side, right=thin_side)
            if row in [12, 18, 26, 48, 24, 31, 34, 46, 51]:
                cell.font = bold_font
            else:
                cell.font = regular_font

    # Scrive le formule esatte utilizzate nella colonna E per documentazione didattica dell'utente
    for r in range(13, 35):
        cell_formula = ws.cell(row=r, column=4).value
        if isinstance(cell_formula, str) and cell_formula.startswith("="):
            ws.cell(row=r, column=5).value = f"'{cell_formula}"
            ws.cell(row=r, column=5).font = italic_font
            ws.cell(row=r, column=5).border = thin_border
            
    # Auto-fit larghezza colonne
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # Evita di allargare a dismisura per la cella A37 che è mergiata
                if cell.coordinate == 'A37':
                    continue
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Salva il file
    output_dir = "data"
    os.makedirs(output_dir, exist_ok=True)
    file_path = os.path.join(output_dir, "Financial_Planning_Model.xlsx")
    wb.save(file_path)
    print(f"[OK] Dynamic Excel Financial Planning Model created at: {os.path.abspath(file_path)}")

if __name__ == "__main__":
    create_planning_model()
