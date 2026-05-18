import sys, subprocess, os

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─── COLOUR PALETTE ───────────────────────────────────────────────────────────
C_DARK   = "1A1A2E"   # dark navy (header bg)
C_BLUE   = "16213E"   # section bg
C_ACCENT = "0F3460"   # sub-header
C_GREEN  = "28A745"   # positive result
C_RED    = "DC3545"   # negative result
C_GOLD   = "E8C84A"   # highlight
C_LIGHT  = "ECF0F1"   # light row fill
C_WHITE  = "FFFFFF"

def hdr_font(size=11, bold=True, color=C_WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def cell_font(size=11, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def style(ws, row, col, value=None, num_format=None,
          bg=None, font=None, align="left", bold=False, color="000000"):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    if bg:
        c.fill = fill(bg)
    c.font = Font(name="Calibri", size=11, bold=bold, color=color)
    if font:
        c.font = font
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = border()
    if num_format:
        c.number_format = num_format
    return c


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – DASHBOARD (Inputs + Results)
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "📊 Dashboard"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 18
ws.row_dimensions[1].height = 40

# ── TITLE ─────────────────────────────────────────────────────────────────────
ws.merge_cells("A1:E1")
c = ws["A1"]
c.value = "CAPM  |  WACC  |  IRR  —  Interactive Financial Simulator"
c.font = Font(name="Calibri", size=16, bold=True, color=C_GOLD)
c.fill = fill(C_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")

# ── SECTION: INPUT PARAMETERS ─────────────────────────────────────────────────
row = 3
ws.merge_cells(f"A{row}:E{row}")
c = ws[f"A{row}"]
c.value = "▶  INPUT PARAMETERS  (yellow cells = change these)"
c.font = hdr_font(12, color=C_GOLD)
c.fill = fill(C_BLUE)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[row].height = 28

inputs = [
    # (label, value, format, note)
    ("Risk-Free Rate  r_f",             0.03,    "0.00%", "e.g. 10yr govt bond yield"),
    ("Market Return  r_m",              0.10,    "0.00%", "e.g. S&P500 avg historical"),
    ("Beta  β",                         1.4,     "0.00",  "Sensitivity to market risk"),
    ("Cost of Debt  r_d",               0.05,    "0.00%", "Bank loan interest rate"),
    ("Corporate Tax Rate  T_c",         0.30,    "0.00%", "Effective corporate tax"),
    ("Market Value of Equity  E (€)",   600000,  "#,##0 €", "Total equity market value"),
    ("Market Value of Debt  D (€)",     400000,  "#,##0 €", "Total debt market value"),
]

# Store which rows hold input values (for formula references later)
INPUT_ROWS = {}
label_to_key = {
    "Risk-Free Rate  r_f": "rf",
    "Market Return  r_m":  "rm",
    "Beta  β":             "beta",
    "Cost of Debt  r_d":   "rd",
    "Corporate Tax Rate  T_c": "tc",
    "Market Value of Equity  E (€)": "E",
    "Market Value of Debt  D (€)":  "D",
}

for i, (label, val, fmt, note) in enumerate(inputs):
    r = row + 1 + i
    ws.row_dimensions[r].height = 22
    style(ws, r, 1, label, bold=False, bg=C_LIGHT)
    c = ws.cell(row=r, column=2)
    c.value = val
    c.number_format = fmt
    c.fill = fill("FFF9C4")          # yellow = editable
    c.font = Font(name="Calibri", size=11, bold=True, color="1A1A2E")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()
    style(ws, r, 3, note, color="555555", bg=C_WHITE)
    INPUT_ROWS[label_to_key[label]] = r

# Named-range shortcuts (row references)
rf   = f"B{INPUT_ROWS['rf']}"
rm   = f"B{INPUT_ROWS['rm']}"
beta = f"B{INPUT_ROWS['beta']}"
rd   = f"B{INPUT_ROWS['rd']}"
tc   = f"B{INPUT_ROWS['tc']}"
E_   = f"B{INPUT_ROWS['E']}"
D_   = f"B{INPUT_ROWS['D']}"
V_   = f"({E_}+{D_})"          # Total value formula string

# ── SECTION: CAPM ─────────────────────────────────────────────────────────────
row = row + len(inputs) + 2
ws.merge_cells(f"A{row}:E{row}")
c = ws[f"A{row}"]
c.value = "📐  STEP 1  —  CAPM  :  Cost of Equity  r_e"
c.font = hdr_font(12, color=C_WHITE)
c.fill = fill(C_ACCENT)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[row].height = 28

capm_rows = [
    ("Formula", "r_e  =  r_f  +  β × (r_m – r_f)", None, "CAPM equation"),
    ("Market Risk Premium  (r_m – r_f)", f"={rm}-{rf}", "0.00%", "Extra return vs risk-free"),
    ("Risk Premium of stock  β×(r_m–r_f)", f"={beta}*({rm}-{rf})", "0.00%", "Stock-specific premium"),
    ("► COST OF EQUITY  r_e", f"={rf}+{beta}*({rm}-{rf})", "0.00%", "Result of CAPM"),
]

CAPM_RESULT_ROW = None
for i, (label, val, fmt, note) in enumerate(capm_rows):
    r = row + 1 + i
    ws.row_dimensions[r].height = 22
    is_result = label.startswith("►")
    bg_row = "E8F5E9" if is_result else C_LIGHT
    style(ws, r, 1, label, bold=is_result, bg=bg_row, color="1A1A2E")
    c = ws.cell(row=r, column=2)
    c.value = val if fmt is None else val
    if fmt:
        c.value = val
        c.number_format = fmt
    c.fill = fill("C8E6C9" if is_result else C_WHITE)
    c.font = Font(name="Calibri", size=11, bold=is_result, color=C_GREEN if is_result else "000000")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()
    style(ws, r, 3, note, color="555555", bg=C_WHITE)
    if is_result:
        CAPM_RESULT_ROW = r

capm_cell = f"B{CAPM_RESULT_ROW}"

# ── SECTION: WACC ─────────────────────────────────────────────────────────────
row = CAPM_RESULT_ROW + 2
ws.merge_cells(f"A{row}:E{row}")
c = ws[f"A{row}"]
c.value = "⚖️  STEP 2  —  WACC  :  Weighted Average Cost of Capital"
c.font = hdr_font(12, color=C_WHITE)
c.fill = fill(C_ACCENT)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[row].height = 28

wacc_rows = [
    ("Formula", "WACC = (E/V)×r_e + (D/V)×r_d×(1–T_c)", None, "Weighted avg of debt & equity cost"),
    ("Total Firm Value  V = E+D", f"={V_}", "#,##0 €", "Sum of equity and debt"),
    ("Weight of Equity  E/V", f"={E_}/{V_}", "0.00%", "Equity fraction"),
    ("Weight of Debt  D/V", f"={D_}/{V_}", "0.00%", "Debt fraction"),
    ("After-tax Cost of Debt  r_d×(1–T_c)", f"={rd}*(1-{tc})", "0.00%", "Debt is tax-deductible (tax shield)"),
    ("► WACC", f"=({E_}/{V_})*{capm_cell}+({D_}/{V_})*{rd}*(1-{tc})", "0.00%", "Hurdle rate for any project"),
]

WACC_RESULT_ROW = None
for i, (label, val, fmt, note) in enumerate(wacc_rows):
    r = row + 1 + i
    ws.row_dimensions[r].height = 22
    is_result = label.startswith("►")
    bg_row = "E3F2FD" if is_result else C_LIGHT
    style(ws, r, 1, label, bold=is_result, bg=bg_row, color="1A1A2E")
    c = ws.cell(row=r, column=2)
    c.value = val if fmt is None else val
    if fmt:
        c.value = val
        c.number_format = fmt
    c.fill = fill("BBDEFB" if is_result else C_WHITE)
    c.font = Font(name="Calibri", size=11, bold=is_result, color="1565C0" if is_result else "000000")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()
    style(ws, r, 3, note, color="555555", bg=C_WHITE)
    if is_result:
        WACC_RESULT_ROW = r

wacc_cell = f"B{WACC_RESULT_ROW}"

# ── SECTION: DECISION ─────────────────────────────────────────────────────────
row = WACC_RESULT_ROW + 2
ws.merge_cells(f"A{row}:E{row}")
c = ws[f"A{row}"]
c.value = "🏆  FINAL DECISION  —  See 'Cash Flows & IRR' sheet for NPV & IRR calculation"
c.font = hdr_font(12, color=C_GOLD)
c.fill = fill(C_DARK)
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[row].height = 28

dec_rows = [
    ("CAPM – Cost of Equity", f"={capm_cell}", "0.00%", "What shareholders demand"),
    ("WACC – Hurdle Rate",     f"={wacc_cell}", "0.00%", "Minimum project return"),
    ("IRR – Project Return",   "=IRR Sheet!B18", "0.00%", "← Link from IRR sheet"),
    ("NPV of Project",         "=IRR Sheet!B19", "#,##0.00 €", "← Link from IRR sheet"),
    ("Decision: IRR > WACC?",
     f'=IF(\'IRR Sheet\'!B18>{wacc_cell},"✅ ACCEPT – Project creates value","❌ REJECT – Destroys value")',
     None, "Golden Rule"),
]

for i, (label, val, fmt, note) in enumerate(dec_rows):
    r = row + 1 + i
    ws.row_dimensions[r].height = 24
    is_last = i == len(dec_rows) - 1
    style(ws, r, 1, label, bold=True, bg="FFF8E1", color="1A1A2E")
    c = ws.cell(row=r, column=2)
    c.value = val
    if fmt:
        c.number_format = fmt
    c.fill = fill("FFECB3" if not is_last else "FFF176")
    c.font = Font(name="Calibri", size=11, bold=True, color="E65100")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border()
    ws.row_dimensions[r].height = 28 if is_last else 24
    style(ws, r, 3, note, color="555555", bg=C_WHITE)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – CASH FLOWS & IRR/NPV
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("IRR Sheet")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 20
ws2.column_dimensions["D"].width = 22

# Title
ws2.merge_cells("A1:D1")
c = ws2["A1"]
c.value = "📈  CASH FLOW ANALYSIS  —  NPV & IRR Calculator"
c.font = Font(name="Calibri", size=14, bold=True, color=C_GOLD)
c.fill = fill(C_DARK)
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 38

# Header row
r = 3
for col, hd in enumerate(["Year", "Cash Flow (€)", "Discount Factor (WACC)", "Present Value (€)"], 1):
    c = ws2.cell(row=r, column=col)
    c.value = hd
    c.font = hdr_font(11, color=C_WHITE)
    c.fill = fill(C_ACCENT)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()
ws2.row_dimensions[r].height = 24

# Cash flow data rows
cf_values = [-100000, 30000, 40000, 50000, 20000]
CF_START = 4
for i, cf in enumerate(cf_values):
    r = CF_START + i
    ws2.row_dimensions[r].height = 22
    year = i
    # Year
    c = ws2.cell(row=r, column=1)
    c.value = f"Year {year}" if year > 0 else "Year 0  (Initial Investment)"
    c.font = Font(name="Calibri", size=11, bold=(year == 0))
    c.fill = fill(C_LIGHT if year % 2 == 0 else C_WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border()

    # Cash Flow (yellow = editable)
    c = ws2.cell(row=r, column=2)
    c.value = cf
    c.number_format = "#,##0 €"
    c.fill = fill("FFF9C4")   # yellow
    c.font = Font(name="Calibri", size=11, bold=True,
                  color=C_RED if cf < 0 else "1A5E20")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()

    # Discount Factor = 1/(1+WACC)^t
    c = ws2.cell(row=r, column=3)
    if year == 0:
        c.value = 1
    else:
        c.value = f"=1/(1+'📊 Dashboard'!{wacc_cell})^{year}"
    c.number_format = "0.0000"
    c.fill = fill(C_LIGHT if year % 2 == 0 else C_WHITE)
    c.font = Font(name="Calibri", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()

    # PV = CF * Discount Factor
    c = ws2.cell(row=r, column=4)
    c.value = f"=B{r}*C{r}"
    c.number_format = "#,##0.00 €"
    c.fill = fill(C_LIGHT if year % 2 == 0 else C_WHITE)
    c.font = Font(name="Calibri", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border()

# Totals / Results
last_cf_row = CF_START + len(cf_values) - 1

result_rows = [
    (last_cf_row + 2, "NPV  (=Sum of all PVs)", f"=SUM(D{CF_START}:D{last_cf_row})",
     "#,##0.00 €", "B18", C_GREEN, "28A745"),
    (last_cf_row + 3, "IRR  (Internal Rate of Return)",
     f"=IRR(B{CF_START}:B{last_cf_row})", "0.00%", "B19", "1565C0", "1565C0"),
    (last_cf_row + 4, "WACC  (Hurdle Rate from Dashboard)",
     f"='📊 Dashboard'!{wacc_cell}", "0.00%", None, "E65100", "E65100"),
    (last_cf_row + 5, "Decision: IRR vs WACC",
     f"=IF(B{last_cf_row+3}>B{last_cf_row+4},\"✅ ACCEPT\",\"❌ REJECT\")",
     None, None, C_GOLD, C_DARK),
]

for (r, label, formula, fmt, _, bg, fc) in result_rows:
    ws2.row_dimensions[r].height = 26
    c = ws2.cell(row=r, column=1)
    c.value = label
    c.font = Font(name="Calibri", size=11, bold=True)
    c.fill = fill(C_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border()

    c = ws2.cell(row=r, column=2)
    c.value = formula
    if fmt:
        c.number_format = fmt
    c.fill = fill(bg)
    c.font = Font(name="Calibri", size=12, bold=True, color=fc)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border()

# Instructions box
ins_row = last_cf_row + 8
ws2.merge_cells(f"A{ins_row}:D{ins_row}")
c = ws2[f"A{ins_row}"]
c.value = ("💡  HOW TO USE: Change the yellow cells (Cash Flows) or return to '📊 Dashboard' "
           "to change Beta, Debt, Equity, Tax Rate etc. All results update automatically.")
c.font = Font(name="Calibri", size=10, italic=True, color="444444")
c.fill = fill("FFFDE7")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws2.row_dimensions[ins_row].height = 40

# ── Save ───────────────────────────────────────────────────────────────────────
out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "wiki", "Progetti", "CAPM_WACC_IRR_Simulator.xlsx")
wb.save(out)
print(f"SUCCESS: {out}")
