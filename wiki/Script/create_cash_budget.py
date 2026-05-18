import sys
import subprocess
import os

try:
    import openpyxl
except ImportError:
    print("openpyxl non trovato. Installazione in corso...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "Short Term Cash Budget"

headers = ["Item", "Month 1", "Month 2", "Month 3", "Month 4", "Total"]
ws.append(headers)

# Formattazione degli header
for col in range(1, 7):
    ws.cell(row=1, column=col).font = Font(bold=True)
    ws.cell(row=1, column=col).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

data = [
    ["Beginning Cash Balance", 10000, "=B30", "=C30", "=D30", ""],
    ["", "", "", "", "", ""],
    ["CASH INFLOWS", "", "", "", "", ""],
    ["Cash Sales", 20000, 24000, 22000, 26000, "=SUM(B5:E5)"],
    ["Collection of Receivables", 15000, 30000, 36000, 33000, "=SUM(B6:E6)"],
    ["Other Inflows", 0, 0, 5000, 0, "=SUM(B7:E7)"],
    ["Total Cash Inflows", "=SUM(B5:B7)", "=SUM(C5:C7)", "=SUM(D5:D7)", "=SUM(E5:E7)", "=SUM(F5:F7)"],
    ["", "", "", "", "", ""],
    ["CASH AVAILABLE", "=B2+B8", "=C2+C8", "=D2+D8", "=E2+E8", ""],
    ["", "", "", "", "", ""],
    ["CASH OUTFLOWS", "", "", "", "", ""],
    ["Payments to Suppliers", 18000, 20000, 21000, 19000, "=SUM(B13:E13)"],
    ["Wages and Salaries", 12000, 12000, 12000, 12000, "=SUM(B14:E14)"],
    ["Rent and Overhead", 4000, 4000, 4000, 4000, "=SUM(B15:E15)"],
    ["Taxes", 0, 3000, 0, 3000, "=SUM(B16:E16)"],
    ["Capital Expenditures", 0, 8000, 10000, 0, "=SUM(B17:E17)"],
    ["Total Cash Outflows", "=SUM(B13:B17)", "=SUM(C13:C17)", "=SUM(D13:D17)", "=SUM(E13:E17)", "=SUM(F13:F17)"],
    ["", "", "", "", "", ""],
    ["NET CASH FLOW", "=B8-B18", "=C8-C18", "=D8-D18", "=E8-E18", "=F8-F18"],
    ["", "", "", "", "", ""],
    ["ENDING CASH BALANCE (Before Fin)", "=B10-B18", "=C10-C18", "=D10-D18", "=E10-E18", ""],
    ["Minimum Cash Requirement", 10000, 10000, 10000, 10000, ""],
    ["Shortfall / Surplus", "=B22-B23", "=C22-C23", "=D22-D23", "=E22-E23", ""],
    ["", "", "", "", "", ""],
    ["FINANCING", "", "", "", "", ""],
    ["Borrowing (If Shortfall < 0)", "=IF(B24<0, -B24, 0)", "=IF(C24<0, -C24, 0)", "=IF(D24<0, -D24, 0)", "=IF(E24<0, -E24, 0)", ""],
    ["Repayments (Manual Input, use negative)", 0, 0, 0, 0, ""],
    ["", "", "", "", "", ""],
    ["ENDING CASH BALANCE (After Fin)", "=B22+B27+B28", "=C22+C27+C28", "=D22+D27+D28", "=E22+E27+E28", ""]
]

for row in data:
    ws.append(row)

# Larghezza colonne
ws.column_dimensions['A'].width = 40
for col in ['B', 'C', 'D', 'E', 'F']:
    ws.column_dimensions[col].width = 15

# Grassetto sulle righe principali
bold_rows = [3, 8, 10, 12, 18, 20, 22, 26, 30]
for r in bold_rows:
    for c in range(1, 7):
        ws.cell(row=r, column=c).font = Font(bold=True)

# Formattazione valuta euro
for r in range(2, 31):
    for c in range(2, 7):
        if ws.cell(row=r, column=c).value != "":
            ws.cell(row=r, column=c).number_format = '#,##0.00 \u20ac'

output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "wiki", "Progetti", "Short_Term_Cash_Budget.xlsx")
wb.save(output_path)
print(f"SUCCESS: File Excel creato in {output_path}")
